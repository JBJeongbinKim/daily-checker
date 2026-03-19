from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from datetime import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE_PATH = ROOT / "data" / "Running.xlsx"
OUTPUT_PATH = ROOT / "data" / "running-program.json"
SHEET_NAME = "Nike training program_21k"


@dataclass(frozen=True)
class DayBlock:
    order: int
    display_column: int
    focus: str


DAY_BLOCKS = [
    DayBlock(order=1, display_column=29, focus="Recovery"),
    DayBlock(order=2, display_column=34, focus="Speed"),
    DayBlock(order=3, display_column=39, focus="Recovery"),
    DayBlock(order=4, display_column=44, focus="Speed"),
    DayBlock(order=5, display_column=49, focus="Long"),
]


def format_clock(value: Any) -> str | None:
    if value in (None, "-"):
        return None

    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"

    if isinstance(value, str):
        return value.strip() or None

    return str(value)


def parse_duration_minutes(value: str | None) -> float:
    if not value:
        return 0

    parts = value.split(":")
    if len(parts) != 2:
        return 0

    minutes, seconds = parts
    return int(minutes) + (int(seconds) / 60)


def round_distance(value: float) -> float:
    return round(value + 1e-9, 2)


def serialize_segment(laps: Any, category: Any, pace: Any, distance: Any, duration: Any) -> dict[str, Any]:
    normalized_pace = format_clock(pace)
    normalized_duration = format_clock(duration)
    distance_km = round_distance(float(distance)) if isinstance(distance, (int, float)) else None
    laps_value = int(laps) if isinstance(laps, (int, float)) and float(laps).is_integer() else laps

    return {
        "laps": laps_value if laps_value not in ("", None) else None,
        "category": category if category not in ("", None, "-") else None,
        "pace": normalized_pace,
        "distanceKm": distance_km,
        "duration": normalized_duration,
        "durationMinutes": round(parse_duration_minutes(normalized_duration), 2) if normalized_duration else 0,
    }


def build_day(rows: list[tuple[Any, Any, Any, Any, Any]], block: DayBlock, week_id: str) -> dict[str, Any]:
    segments = [serialize_segment(*row) for row in rows]
    total_distance = round_distance(sum(segment["distanceKm"] or 0 for segment in segments))
    total_duration = round(sum(segment["durationMinutes"] for segment in segments), 2)

    paces = [segment["pace"] for segment in segments if segment["pace"]]
    unique_paces = []
    for pace in paces:
        if pace not in unique_paces:
            unique_paces.append(pace)

    if not unique_paces:
        pace_summary = "Not listed"
    elif len(unique_paces) == 1:
        pace_summary = f"{unique_paces[0]}/km"
    else:
        pace_summary = "Mixed"

    primary_metric = {
        "kind": "distance" if total_distance > 0 else "duration",
        "distanceKm": total_distance if total_distance > 0 else None,
        "durationMinutes": total_duration if total_distance == 0 else None,
    }

    return {
        "id": f"{week_id}-day-{block.order}",
        "order": block.order,
        "label": f"Day {block.order}",
        "focus": block.focus,
        "paceSummary": pace_summary,
        "totalDistanceKm": total_distance,
        "totalDurationMinutes": total_duration,
        "primaryMetric": primary_metric,
        "segments": segments,
    }


def main() -> None:
    workbook = load_workbook(SOURCE_PATH, data_only=True)
    sheet = workbook[SHEET_NAME]

    week_rows: list[int] = []
    for row_index in range(1, sheet.max_row + 1):
        week_value = sheet.cell(row_index, 2).value
        if isinstance(week_value, str) and week_value.startswith("W-"):
            week_rows.append(row_index)

    boundaries = week_rows + [sheet.max_row + 1]
    weeks = []

    for index, start_row in enumerate(week_rows):
        end_row = boundaries[index + 1]
        week_id = str(sheet.cell(start_row, 2).value)
        days = []

        for block in DAY_BLOCKS:
            block_rows: list[tuple[Any, Any, Any, Any, Any]] = []
            for row_index in range(start_row, end_row):
                values = tuple(sheet.cell(row_index, block.display_column + offset).value for offset in range(5))
                if any(value not in (None, "-", "") for value in values):
                    block_rows.append(values)

            days.append(build_day(block_rows, block, week_id))

        weeks.append(
            {
                "id": week_id,
                "label": f"Week {week_id.removeprefix('W-')}",
                "order": index,
                "days": days,
            }
        )

    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": str(SOURCE_PATH.relative_to(ROOT)).replace("\\", "/"),
        "sheetName": SHEET_NAME,
        "weekCount": len(weeks),
        "weeks": weeks,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {len(weeks)} weeks to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
