from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE_PATH = ROOT / "data" / "Rent.xlsx"
OUTPUT_PATH = ROOT / "data" / "rent-history.json"
SOURCE_URL = "https://verisresidential.com/jersey-city-nj-apartments/the-blvd-collection/"


def to_iso(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def to_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "unit"


def get_header_date(worksheet, column_index: int) -> str | None:
    for row_index in (3, 5, 2, 4):
        value = worksheet.cell(row=row_index, column=column_index).value
        iso_value = to_iso(value)
        if iso_value:
            return iso_value
    return None


def main() -> None:
    workbook = load_workbook(SOURCE_PATH, data_only=True)
    worksheet = workbook["Sheet1"]

    price_columns: list[tuple[int, str]] = []
    for column_index in range(5, worksheet.max_column + 1):
        snapshot_date = get_header_date(worksheet, column_index)
        if not snapshot_date:
            continue

        has_price = any(
            isinstance(worksheet.cell(row=row_index, column=column_index).value, (int, float))
            for row_index in range(6, worksheet.max_row + 1)
        )

        if has_price:
            price_columns.append((column_index, snapshot_date))

    units = []
    all_dates = sorted({snapshot_date for _, snapshot_date in price_columns})
    latest_snapshot_date = all_dates[-1] if all_dates else None

    for row_index in range(6, worksheet.max_row + 1):
        raw_label = worksheet.cell(row=row_index, column=2).value
        unit_number = worksheet.cell(row=row_index, column=3).value
        availability_value = worksheet.cell(row=row_index, column=4).value

        if not raw_label and not unit_number:
            continue

        label = str(raw_label or "").strip()
        unit = str(unit_number or "").strip()

        if not label or not unit:
            continue

        parts = label.split()
        building_id = parts[0]
        layout_id = parts[1] if len(parts) > 1 else label

        snapshots_by_date: dict[str, int] = {}
        for column_index, snapshot_date in price_columns:
            price_value = worksheet.cell(row=row_index, column=column_index).value
            if isinstance(price_value, (int, float)):
                snapshots_by_date[snapshot_date] = int(round(price_value))

        snapshots = [
            {
                "date": snapshot_date,
                "price": snapshots_by_date[snapshot_date],
                "isThursday": datetime.fromisoformat(snapshot_date).weekday() == 3,
            }
            for snapshot_date in sorted(snapshots_by_date.keys())
        ]

        first_seen = snapshots[0]["date"] if snapshots else None
        last_seen = snapshots[-1]["date"] if snapshots else None
        current_price = snapshots[-1]["price"] if snapshots else None
        initial_price = snapshots[0]["price"] if snapshots else None
        change_since_first = (
            current_price - initial_price if current_price is not None and initial_price is not None else None
        )

        units.append(
            {
                "id": to_slug(f"{building_id}-{unit}"),
                "buildingId": building_id,
                "layoutId": layout_id,
                "typeLabel": label,
                "unitNumber": unit,
                "availabilityDate": to_iso(availability_value),
                "status": "active" if last_seen and latest_snapshot_date and last_seen == latest_snapshot_date else "inactive",
                "firstSeen": first_seen,
                "lastSeen": last_seen,
                "currentPrice": current_price,
                "initialPrice": initial_price,
                "changeSinceFirst": change_since_first,
                "snapshotCount": len(snapshots),
                "snapshots": snapshots,
            }
        )

    units.sort(key=lambda unit: (unit["buildingId"], unit["layoutId"], unit["unitNumber"]))

    dataset = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": str(SOURCE_PATH.relative_to(ROOT)).replace("\\", "/"),
        "sourceUrl": SOURCE_URL,
        "latestSnapshotDate": latest_snapshot_date,
        "snapshotDates": all_dates,
        "totals": {
            "units": len(units),
            "activeUnits": sum(1 for unit in units if unit["status"] == "active"),
            "inactiveUnits": sum(1 for unit in units if unit["status"] == "inactive"),
            "snapshots": sum(unit["snapshotCount"] for unit in units),
        },
        "units": units,
    }

    OUTPUT_PATH.write_text(json.dumps(dataset, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
